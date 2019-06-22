from openpyxl import Workbook
from openpyxl import load_workbook
# wb = Workbook()


# ws = wb.active
# ws.title = "Customer"
# wb.create_sheet("Invoices")
# wb.create_sheet("Invoice_Line_Items")
# wb.create_sheet("Product")

# for sheet in wb:
# 	sheet.sheet_properties.tabColor = "1072BA"

# print(wb.sheetnames)

# for sheet in wb:
# 	print(sheet.title)


# wb["Customer"]['A4'] = 4
# wb.save('Comp_230.xlsx')


# wb_invoice = load_workbook('invoices.xlsx')
# print (wb_invoice.sheetnames)


# #wb_invoice["cust_name"] = 

# for row in wb_invoice["customers"].iter_rows(min_row=1, max_col=2, max_row=20,values_only=True):
#    for cell in row:
#    	print(cell)


wb = load_workbook('invoices.xlsx')
print(wb.sheetnames)

cust_ws = wb["customers"]
inv_ws = wb["invoices"]
lines_ws = wb["line_items"]
prod_ws = wb["products"]

#invnum = int(input('Enter invoice Number to generate invoice: '))
invnum = 70

for row in inv_ws.iter_rows(min_row=2, max_col=3, values_only=True):
    if row[0] == invnum:
        cust_ID = row[1]
        invoice_date = row[2]

for row in cust_ws.iter_rows(min_row=2, max_col=2, values_only=True):
    if row[0] == cust_ID:
        cust_name = row[1]

lines_list = []
tot_inv_cost = 0

for row in lines_ws.iter_rows(min_row=2, max_col=3, values_only=True):
    if row[0] == invnum:
        for rowprod in prod_ws.iter_rows(min_row=2, values_only=True):
            if rowprod[0] == row[1]:
                sub_total = row[2] * rowprod[2]
                lines_list.append({"prod_id": row[1], "prod_name": rowprod[1], "quantity": row[2], "prod_cost": rowprod[2], "subtotal": sub_total})
                tot_inv_cost += sub_total

invoice_output = load_workbook("invoice_template.xlsx")

ws = invoice_output["Invoice"]

ws['B3'] = invnum
ws["B4"] = invoice_date
ws["B5"] = cust_name
ws["B6"] = cust_ID
ws["E18"] = tot_inv_cost

for i, line in enumerate(lines_list):
    ws.cell(row=i+9, column=1, value=line["prod_name"])
    ws.cell(row=i+9, column=2, value=line["prod_id"])
    ws.cell(row=i+9, column=3, value=line["prod_cost"])
    ws.cell(row=i+9, column=4, value=line["quantity"])
    ws.cell(row=i+9, column=5, value=line["subtotal"])
    # print("here",line)
    # print(lines_list)

# newfilename = "invoice" + str(invnum) + ".xlsx"
# invoice_output.save(newfilename)






class Basic_obj(dict):
   def __getattr__(self, item):
       return self[item]


def key_value_obj_dict(header_key_value, object_key_value):
	# length = len(header_key_value)
	cust_obj = Basic_obj()
	for i in range(len(header_key_value)):
		cust_obj[header_key_value[i]] = object_key_value[i]
	# print("test",cust_obj)
	return(cust_obj)


def create_dict(sheet):
	cust_dict={}
	counter = 0
	header=""
	dup_key=[]
	none_key=[]
	none_key_value=[]
	line = 0
	for row in sheet.values:
		if counter == 0:
			header_key_value = list(row)
			header_key = header_key_value.pop(0)
			counter = counter+1
		else:
			object_key_value = list(row)
			cust_key = object_key_value.pop(0)
			the_value = key_value_obj_dict(header_key_value, object_key_value)
			# print(cust_key)
			if cust_key in cust_dict:
				dup_key.append([cust_key,the_value])
			elif cust_key is None:
				none_key.append([cust_key,the_value])
			elif the_value is None:
				none_key_value.append([cust_key,the_value])
			else:
				cust_dict[cust_key] = the_value
		line = line + 1
	print("dup_key test", dup_key)
	print("none_key test", none_key)
	print("none_key_value test", none_key)
	print("Number of lines in the sheet:", line)
	print("Number of Unique Customers", len(cust_dict))
	print("Number of entries in the sheet",len(cust_dict) + len(dup_key))
	# cust_dict = [cust_dict,dup_key]
	# for item in cust_dict:
	# 	print (f"{item}:{cust_dict[item]}")
	return cust_dict
			



class Customer():
	def __init__(self,cust_id,cust_name):
		self.cust_id = cust_id
		self.cust_name = cust_name

class Product():
	def __init__(self,prod_id, prod_name, prod_cost,inventory):
		self.prod_id = prod_id
		self.prod_name = prod_name
		self.prod_cost = prod_cost
		self.inventory = inventory

class Invoice():
	def __init__(self, inv_id, cust_id,invoice_date):
		self.inv_id = inv_id
		self.cust_id = cust_id
		self.invoice_date = invoice_date
		prod_details = ""



class Helper():
	def __init__(self,book):
		self.book = book
		from openpyxl import Workbook
		from openpyxl import load_workbook
		self.wb = load_workbook(book)
		self.cust_ws = wb["customers"]
		self.inv_ws = wb["invoices"]
		self.lines_ws = wb["line_items"]
		self.prod_ws = wb["products"]

	def cust_dict(self):
		self.cust_dict = {}
		self.cust_dup = []
		self.none_key=[]
		self.none_obj=[]
		for row in self.cust_ws.iter_rows(min_row=2,values_only=True):
			self.key = row[0]
			self.obj = Customer(row[0],row[1])
			if self.key in self.cust_dict:
				self.cust_dup.extend([self.key,self.obj])
			elif self.key is None:
				self.none_key.extend([self.key, self.obj])
			elif self.obj.cust_name is None:
				self.none_obj.extend([self.key,self.obj])
			else:
				self.cust_dict[self.key] = self.obj

		print("Duplicate Customers Items",self.cust_dup)
		print("Missing Customers Key Data",self.none_key)
		print("Missing Customers Key Value Data",self.none_obj)	
		# print(self.cust_dict)
		# print("print from line 240", self.cust_dict[7].cust_name)
		return self.cust_dict
		# print("print from line 240", self.cust_dict[7].cust_name)

	def prod_dict(self):
		self.prod_dict = {}
		self.prod_dup = []
		self.none_key=[]
		self.none_obj=[]
		for row in self.prod_ws.iter_rows(min_row=2,values_only=True):
			self.key = row[0]
			self.obj = Product(row[0],row[1],row[2], row[3])
			if self.key in self.prod_dict:
				self.prod_dup.extend([self.key,self.obj])
			elif self.key is None:
				self.none_key.extend([self.key, self.obj])
			elif self.obj.prod_name is None:
				self.none_obj.extend([self.key,self.obj])
			else:
				self.prod_dict[self.key] = self.obj

		print("Duplicate Product Items",self.prod_dup)
		print("Missing Product Key Data",self.none_key)
		print("Missing Product Key Value Data",self.none_obj)	
		print(self.prod_dict)
		# print("print from line 240", self.cust_dict[7].cust_name)
		return self.prod_dict


	def item(self, invoiceid):
		self.item_dict = {}
		self.none_list = []
		for row in self.lines_ws.iter_rows(min_row=2, values_only=True):
			if row[0] != None and row[1] != None and row[2] != None:
				if row[0] == invoiceid:
					self.item_dict[row[1]] = (row[2])
			else:
				self.none_list.append({row[0]: {row[1]: row[2]}})
		# print('Missing Data', self.none_list)

		return self.item_dict

	def inv_dict(self):
		self.inv_dict = {}
		self.inv_dup = []
		self.none_key=[]
		self.none_obj=[]

		for row in self.inv_ws.iter_rows(min_row=2, values_only=True):
			self.key = row[0]
			self.obj = Invoice(row[0],row[1],row[2])
			if self.key not in self.inv_ws and self.key != None and self.obj!= None:
				self.inv_dict[self.key] = self.obj
				self.inv_dict[self.key].prod_details = self.item(self.key)
			else:
				self.inv_dup.extend({self.key:self.obj})
		print(self.inv_dict[55].prod_details)
		# print(self.inv_dict)return self.inv_dict

	def make_invoice(self, invoiceid):
		from openpyxl import load_workbook
		invoice_output = load_workbook('Invoice_template.xlsx')
		ws = invoice_output['Invoice']

		self.cust_dict()
		self.prod_dict()
		# self.inv_dict()
		
		# self.cust_dict()
		# self.prod_dict()
		# self.inv_dict()

		b = self.inv_dict[invoiceid].cust_id

		ws['B3'] = invoiceid
		ws['B4'] = self.inv_dict[invoiceid].invoice_date
		ws['B5'] = self.cust_dict[b].cust_name
		ws['B6'] = self.inv_dict[invoiceid].cust_id

		c = self.inv_dict[invoiceid].prod_details
		prod_list = []
		count = 0
		inv_total=0
		for i in c:
			pd_name = self.prod_dict[i].prod_name
			pd_id = i
			pd_price = self.prod_dict[i].prod_cost
			pd_quantity= c[i]
			total_pd_cost = pd_price * pd_quantity
			inv_total = inv_total + total_pd_cost
			prod_list.extend([[pd_name, pd_id,pd_price,pd_quantity,total_pd_cost]])
			count = count+1
		print("test",prod_list)

		for i, line in enumerate(prod_list):
			ws.cell(row=i+9, column=1,value=line[0])
			ws.cell(row=i+9, column=2,value=line[1])
			ws.cell(row=i+9, column=3,value=line[2])
			ws.cell(row=i+9, column=4,value=line[3])
			ws.cell(row=i+9, column=5,value=line[4])
		ws[f"D{10+count}"] = "Total"
		ws[f"E{10+count}"] = inv_total

		newfilename = f"invoice {str(invoiceid)}.xlsx"
		invoice_output.save(newfilename)



test = Helper("invoices.xlsx")
# test.cust_dict()
# test.prod_dict()
# print("test",test.item(51))
test.inv_dict()
test.make_invoice(51)
# print("print from line 288",test.cust_dict[7].cust_name)
# print("print from line 289",test.prod_dict[6].prod_name)

x = ["hello", "gracias", "bye", "goodbye"]
y = [2.65, 8.76, 1.75, 4.26]

# x = ["a", "b", "c", "d", "e", "f", "g", "h", "i"]
# y = [ 0,   1,   1,    0,   1,   2,   2,   0,   1]


###Doing one step at a time to understand more clearly
xy = list(zip(y,x))
print(xy)
xy.sort()
print(xy)
x_sorted = [x for y, x in xy]
print(x_sorted)

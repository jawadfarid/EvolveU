from flask import Flask, render_template, jsonify
from flask_restful import Resource, Api
from flask_cors import CORS

app = Flask(__name__)
# api= Api(app)
CORS(app)



# def cust_dict(book):
# 	from openpyxl import Workbook
# 	from openpyxl import load_workbook
# 	wb = load_workbook(book)
# 	ws = wb.active
# 	cust_dict = {}
# 	cust_dup = []
# 	none_key=[]
# 	none_obj=[]
# 	for row in ws.iter_rows(min_row=2,values_only=True):
# 		key = row[0]
# 		obj = row[1]
# 		if key in cust_dict:
# 			cust_dup.extend([key,obj])
# 		elif key is None:
# 			none_key.extend([key, obj])
# 		elif obj is None:
# 			none_obj.extend([key,obj])
# 		else:
# 			cust_dict[key] = obj

	# print("Duplicate Customers Items",cust_dup)
	# print("Missing Customers Key Data",none_key)
	# print("Missing Customers Key Value Data",none_obj)	
	# # print(cust_dict)
	# # print("print from line 240", cust_dict[7].cust_name)
class Customer():
	def __init__(self,cust_id,cust_name):
		self.cust_id = cust_id
		self.cust_name = cust_name

	# def show(self):
	# 	print(self.__dict__)

	def to_dict(self):
		test_dict={}
		# return {test_dict[self.cust_id]: self.cust_name }
		return { "id" : self.cust_id,
				"name" : self.cust_name}


class Product():
	def __init__(self,prod_id, prod_name, prod_cost,inventory):
		self.prod_id = prod_id
		self.prod_name = prod_name
		self.prod_cost = prod_cost
		self.inventory = inventory

class Invoice():
	def __init__(self, inv_id, cust_id,invoice_date):
		self.inv_id = inv_id
		self.cust_id = cust_id
		self.invoice_date = invoice_date
		prod_details = ""



class Helper():
	def __init__(self,book):
		self.book = book
		from openpyxl import Workbook
		from openpyxl import load_workbook
		self.wb = load_workbook(book)
		self.cust_ws = self.wb["customers"]
		self.inv_ws = self.wb["invoices"]
		self.lines_ws = self.wb["line_items"]
		self.prod_ws = self.wb["products"]

	def cust_dict(self):
		self.cust_dict = {}
		self.cust_dup = []
		self.none_key=[]
		self.none_obj=[]
		for row in self.cust_ws.iter_rows(min_row=2,values_only=True):
			self.key = row[0]
			self.obj = Customer(row[0],row[1])
			if self.key in self.cust_dict:
				self.cust_dup.extend([self.key,self.obj])
			elif self.key is None:
				self.none_key.extend([self.key, self.obj])
			elif self.obj.cust_name is None:
				self.none_obj.extend([self.key,self.obj])
			else:
				self.cust_dict[self.key] = self.obj

		print("Duplicate Customers Items",self.cust_dup)
		print("Missing Customers Key Data",self.none_key)
		print("Missing Customers Key Value Data",self.none_obj)	
		# print(self.cust_dict)
		# print("print from line 240", self.cust_dict[7].cust_name)
		return self.cust_dict
		# print("print from line 240", self.cust_dict[7].cust_name)

	def prod_dict(self):
		self.prod_dict = {}
		self.prod_dup = []
		self.none_key=[]
		self.none_obj=[]
		for row in self.prod_ws.iter_rows(min_row=2,values_only=True):
			self.key = row[0]
			self.obj = Product(row[0],row[1],row[2], row[3])
			if self.key in self.prod_dict:
				self.prod_dup.extend([self.key,self.obj])
			elif self.key is None:
				self.none_key.extend([self.key, self.obj])
			elif self.obj.prod_name is None:
				self.none_obj.extend([self.key,self.obj])
			else:
				self.prod_dict[self.key] = self.obj

		print("Duplicate Product Items",self.prod_dup)
		print("Missing Product Key Data",self.none_key)
		print("Missing Product Key Value Data",self.none_obj)	
		print(self.prod_dict)
		# print("print from line 240", self.cust_dict[7].cust_name)
		return self.prod_dict


	def item(self, invoiceid):
		self.item_dict = {}
		self.none_list = []
		for row in self.lines_ws.iter_rows(min_row=2, values_only=True):
			if row[0] != None and row[1] != None and row[2] != None:
				if row[0] == invoiceid:
					self.item_dict[row[1]] = (row[2])
			else:
				self.none_list.append({row[0]: {row[1]: row[2]}})
		# print('Missing Data', self.none_list)

		return self.item_dict

	def inv_dict(self):
		self.inv_dict = {}
		self.inv_dup = []
		self.none_key=[]
		self.none_obj=[]

		for row in self.inv_ws.iter_rows(min_row=2, values_only=True):
			self.key = row[0]
			self.obj = Invoice(row[0],row[1],row[2])
			if self.key not in self.inv_ws and self.key != None and self.obj!= None:
				self.inv_dict[self.key] = self.obj
				self.inv_dict[self.key].prod_details = self.item(self.key)
			else:
				self.inv_dup.extend({self.key:self.obj})
		print(self.inv_dict[55].prod_details)


@app.route("/")
def info():
	test = Helper("invoices.xlsx")
	test_dict = test.cust_dict()
	return str(test_dict)

@app.route("/customer")
def customer():
	test = Helper("invoices.xlsx")
	test_dict = test.cust_dict()
	return render_template("index.html",test_dict=test_dict)


# my_list = []

# class Comp240(Resource):
# 	def get(self):
# 		test = Helper("invoices.xlsx")
# 		test_dict = test.cust_dict()
# 		# my_list = []
# 		for key in test_dict:
# 			my_list.append(test_dict[key].to_dict())
# 		return jsonify(my_list)


# class Customer_JSON(Resource):
# 	def get(self,id):
# 		test = Helper("invoices.xlsx")
# 		test_dict = test.cust_dict()
		# my_list = []
		# for key in my_list:
		# 	if key["id"] == int(id):
		# 		return key
		
		# return {"id": None}			
		# return my_list	
			
	
	# def post(self,id):
		# cust = {"cust_id":id}
		# my_list.append(cust)
		# return cust
	

	# def delete(self,id):
		# for ind,cust in enumerate(my_list):
		# 	if cust["id"] == id:
		# 		deleted_cust = my_list.pop(ind)
		# 		return {"note":"delete success"}

			
@app.route("/customer_list")
def customer2():
	my_list = []
	test = Helper("invoices.xlsx")
	test_dict = test.cust_dict()
	for key in test_dict:
		my_list.append(test_dict[key].to_dict())
	return jsonify(my_list)

# api.add_resource(Comp240,"/api")
# api.add_resource(Customer_JSON,"/customer/<string:id>")

if __name__ == '__main__':
	app.run(debug=True)